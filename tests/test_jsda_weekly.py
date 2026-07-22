import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from collector import jsda_weekly


FIXTURES = Path(__file__).resolve().parent / "fixtures"
Z_FIXTURE = FIXTURES / "20260710z_sample.xlsx"
S_FIXTURE = FIXTURES / "20260710s_sample.xlsx"


class JSDAWeeklyTests(unittest.TestCase):
    def _modified_copy(self, source, mutate):
        tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(tempdir.cleanup)
        destination = Path(tempdir.name) / source.name
        shutil.copyfile(source, destination)
        workbook = load_workbook(destination)
        mutate(workbook[workbook.sheetnames[0]])
        workbook.save(destination)
        workbook.close()
        return destination

    def test_parse_zandaka_real_values_and_normalization(self):
        issues = jsda_weekly.parse_zandaka(Z_FIXTURE)

        kioxia = issues["285A"]["taishaku"]
        self.assertEqual(kioxia["yutanpo"]["lend_qty"], 9_215_104)
        self.assertEqual(kioxia["mutanpo"]["lend_qty"], 1_393_235)
        self.assertIn("1301", issues)
        self.assertEqual(issues["1301"]["name"], "極 洋")
        self.assertIn("2560", issues)  # The fixture row contains '-' in week-on-week columns.
        self.assertIsNone(
            jsda_weekly._to_int_or_none("-", allow_none=True, context="test")
        )

    def test_normalize_code_keeps_class_share_suffix(self):
        # 末尾'0'(普通株)のみ4桁化し、優先株・社債型種類株式(末尾5等)は5桁のまま
        self.assertEqual(jsda_weekly._normalize_code("13010"), "1301")
        self.assertEqual(jsda_weekly._normalize_code(13010), "1301")
        self.assertEqual(jsda_weekly._normalize_code("285A0"), "285A")
        self.assertEqual(jsda_weekly._normalize_code("25935"), "25935")
        self.assertEqual(jsda_weekly._normalize_code("94346"), "94346")
        with self.assertRaises(ValueError):
            jsda_weekly._normalize_code("")

    def test_parse_shinki_and_build_weekly(self):
        shinki = jsda_weekly.parse_shinki(S_FIXTURE)
        self.assertIn("285A", shinki)

        weekly = jsda_weekly.build_weekly(Z_FIXTURE, S_FIXTURE, "2026-07-10")
        self.assertEqual(weekly["schema_version"], "supply_demand_weekly_v1")
        self.assertEqual(weekly["report_date"], "2026-07-10")
        self.assertEqual(
            weekly["source_files"],
            [Z_FIXTURE.name, S_FIXTURE.name],
        )
        self.assertIn("taishaku", weekly["issues"]["285A"])
        self.assertIn("shinki", weekly["issues"]["285A"])

    def test_all_six_columns_mapped_correctly(self):
        # 極洋(実ファイル8行目由来)の6値を実測値と照合し、列オフセットずれを検知する
        issues = jsda_weekly.parse_zandaka(Z_FIXTURE)
        yutanpo = issues["1301"]["taishaku"]["yutanpo"]
        self.assertEqual(yutanpo["lend_qty"], 103_746)
        self.assertEqual(yutanpo["lend_amt"], 467)
        self.assertEqual(yutanpo["own_qty"], 55_407)
        self.assertEqual(yutanpo["own_amt"], 249)
        self.assertEqual(yutanpo["ten_qty"], 109_771)
        self.assertEqual(yutanpo["ten_amt"], 494)

    def test_invalid_collateral_raises(self):
        path = self._modified_copy(Z_FIXTURE, lambda sheet: setattr(sheet["C8"], "value", "不明"))
        with self.assertRaisesRegex(ValueError, "担保区分"):
            jsda_weekly.parse_zandaka(path)

    def test_normalize_code_accepts_xls_float(self):
        # 旧.xls(xlrd)はコードもfloatで返る(13010.0)
        self.assertEqual(jsda_weekly._normalize_code(13010.0), "1301")
        self.assertEqual(jsda_weekly._normalize_code(25935.0), "25935")

    def test_filename_date_mismatch_raises(self):
        with self.assertRaisesRegex(jsda_weekly.JSDAParseError, "一致しません"):
            jsda_weekly.build_weekly(Z_FIXTURE, S_FIXTURE, "2026-07-03")

    def test_filename_kind_mismatch_raises(self):
        # z引数にsファイルを渡す取り違えを検知する
        with self.assertRaisesRegex(jsda_weekly.JSDAParseError, "種別"):
            jsda_weekly.build_weekly(S_FIXTURE, Z_FIXTURE, "2026-07-10")

    def test_duplicate_collateral_row_raises(self):
        def duplicate_row(sheet):
            values = [sheet.cell(8, c).value for c in range(1, 16)]
            sheet.insert_rows(9)
            for c, v in enumerate(values, 1):
                sheet.cell(9, c).value = v

        path = self._modified_copy(Z_FIXTURE, duplicate_row)
        with self.assertRaisesRegex(jsda_weekly.JSDAParseError, "重複"):
            jsda_weekly.parse_zandaka(path)

    def test_empty_numeric_cell_raises(self):
        path = self._modified_copy(Z_FIXTURE, lambda sheet: setattr(sheet["E8"], "value", None))
        with self.assertRaisesRegex(jsda_weekly.JSDAParseError, "空です"):
            jsda_weekly.parse_zandaka(path)

    def test_header_change_raises(self):
        path = self._modified_copy(Z_FIXTURE, lambda sheet: setattr(sheet["A6"], "value", "変更"))
        with self.assertRaises(jsda_weekly.JSDAParseError):
            jsda_weekly.parse_zandaka(path)

    def test_total_mismatch_raises(self):
        def change_total(sheet):
            sheet.cell(sheet.max_row, 4).value += 1

        path = self._modified_copy(Z_FIXTURE, change_total)
        with self.assertRaises(ValueError):
            jsda_weekly.parse_zandaka(path)

    def test_empty_sheet_raises(self):
        def remove_data(sheet):
            if sheet.max_row >= 8:
                sheet.delete_rows(8, sheet.max_row - 7)

        path = self._modified_copy(Z_FIXTURE, remove_data)
        with self.assertRaisesRegex(ValueError, "データ行が0件"):
            jsda_weekly.parse_zandaka(path)


if __name__ == "__main__":
    unittest.main()
