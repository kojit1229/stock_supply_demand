"""Parse JSDA weekly stock lending workbooks.

File contract (approved by the supervisor; do not change): the weekly output is
``{schema_version, report_date, source_files, issues}``, and each issue stores
the six raw quantity/amount values defined in design.md section 4.
"""

from __future__ import annotations

import argparse
from contextlib import contextmanager
import json
import os
import re
import shutil
import sys
import time
from datetime import date
from pathlib import Path
from typing import Any, Iterator
from urllib.request import Request, urlopen

from openpyxl import load_workbook


SCHEMA_VERSION = "supply_demand_weekly_v1"
USER_AGENT = "stock_supply_demand/1.0 (+https://github.com/kojit1229/stock_supply_demand)"

_ZANDAKA_SHEET = "残高（株券等（上場））"
_SHINKI_SHEET = "新規（株券等（上場））"
_HEADER_ROW_7 = (
    None,
    None,
    None,
    "数量",
    "前週比",
    "金額",
    "前週比",
    "数量",
    "前週比",
    "金額",
    "前週比",
    "数量",
    "前週比",
    "金額",
    "前週比",
)
_HEADER_ROW_6 = {
    "taishaku": (
        "銘柄名",
        "コード",
        "担保",
        "貸付残高",
        None,
        None,
        None,
        "借入残高（自己）",
        None,
        None,
        None,
        "借入残高（転貸）",
        None,
        None,
        None,
    ),
    "shinki": (
        "銘柄名",
        "コード",
        "担保",
        "新規貸付成約高",
        None,
        None,
        None,
        "新規借入成約高（自己）",
        None,
        None,
        None,
        "新規借入成約高（転貸）",
        None,
        None,
        None,
    ),
}
_SHEETS = {"taishaku": _ZANDAKA_SHEET, "shinki": _SHINKI_SHEET}
_COLLATERAL_KEYS = {"有担保": "yutanpo", "無担保": "mutanpo"}
_OUTPUT_COLUMNS = {
    "lend_qty": 3,
    "lend_amt": 5,
    "own_qty": 7,
    "own_amt": 9,
    "ten_qty": 11,
    "ten_amt": 13,
}
_QUANTITY_COLUMNS = (3, 7, 11)
_NUMERIC_COLUMNS = tuple(range(3, 15))
_CHANGE_COLUMNS = (4, 6, 8, 10, 12, 14)
_OLE_MAGIC = bytes.fromhex("D0CF11E0A1B11AE1")


class JSDAParseError(ValueError):
    """Raised when a workbook violates the audited JSDA format."""


def _normalize_name(value: Any) -> str:
    if value is None:
        raise JSDAParseError("銘柄名が空です")
    name = re.sub(r"\s+", " ", str(value).strip())
    if not name:
        raise JSDAParseError("銘柄名が空です")
    return name


def _normalize_code(value: Any) -> str:
    if value is None or not str(value).strip():
        raise JSDAParseError("コードが空です")
    # 旧.xls(xlrd)は数値セルをfloatで返す(13010.0)ため、整数値は先にintへ落とす
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    code = str(value).strip()
    # 統一コード5桁の末尾'0'は普通株の付加桁なので落とす。末尾が'0'以外
    # (優先株・社債型種類株式など、例 25935)は独立銘柄なので5桁のまま保持する
    if len(code) == 5 and code.endswith("0"):
        code = code[:-1]
    if not re.fullmatch(r"[0-9A-Z]{4,5}", code):
        raise JSDAParseError(f"不正な銘柄コードです: {value!r}")
    return code


def _to_int_or_none(value: Any, *, allow_none: bool, context: str) -> int | None:
    """Convert an Excel number to int; a previous-week '-' becomes None."""
    if value == "-":
        if allow_none:
            return None
        raise JSDAParseError(f"{context}に'-'は使用できません")
    if value is None:
        # 監査済みの欠損表記は前週比列の'-'のみ。空セルは欠損ではなく破損として扱う
        raise JSDAParseError(f"{context}が空です")
    if isinstance(value, bool):
        raise JSDAParseError(f"{context}が数値ではありません: {value!r}")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    raise JSDAParseError(f"{context}が整数ではありません: {value!r}")


def _row_tuple(row: tuple[Any, ...]) -> tuple[Any, ...]:
    """Return all 15 audited columns, preserving extra columns for detection."""
    # xlrd represents empty BIFF cells as "" while openpyxl returns None.
    row = tuple(None if value == "" else value for value in row)
    if len(row) >= 15:
        return tuple(row)
    return tuple(row) + (None,) * (15 - len(row))


def _validate_header(row_number: int, row: tuple[Any, ...], kind: str) -> None:
    expected = _HEADER_ROW_6[kind] if row_number == 6 else _HEADER_ROW_7
    actual = _row_tuple(row)
    if actual != expected:
        raise JSDAParseError(
            f"{row_number}行目のヘッダが期待値と一致しません: "
            f"expected={expected!r}, actual={actual!r}"
        )


def _workbook_format(path: str | os.PathLike[str]) -> str:
    """Return the Excel engine selected from an audited extension and magic bytes."""
    workbook_path = Path(path)
    if workbook_path.suffix.lower() not in {".xls", ".xlsx"}:
        raise JSDAParseError(f"Excel拡張子が.xls/.xlsxではありません: {workbook_path}")
    try:
        with workbook_path.open("rb") as source:
            magic = source.read(8)
    except OSError as exc:
        raise JSDAParseError(f"Excelファイルを読めません: {workbook_path}") from exc
    if magic.startswith(b"PK"):
        return "openpyxl"
    if magic == _OLE_MAGIC:
        return "xlrd"
    raise JSDAParseError(
        f"ExcelマジックバイトがPK/OLEではありません: {workbook_path} ({magic.hex()})"
    )


@contextmanager
def _workbook_rows(
    path: str | os.PathLike[str], kind: str
) -> Iterator[Iterator[tuple[Any, ...]]]:
    """Yield audited sheet rows from either OOXML or BIFF8 without duplicating parsing."""
    workbook_path = Path(path)
    engine = _workbook_format(workbook_path)
    sheet_name = _SHEETS[kind]
    if engine == "openpyxl":
        source = None
        workbook = None
        try:
            source = workbook_path.open("rb")
            workbook = load_workbook(source, read_only=True, data_only=True)
            if sheet_name not in workbook.sheetnames:
                raise JSDAParseError(f"対象シートがありません: {sheet_name}")
            yield workbook[sheet_name].iter_rows(values_only=True)
        except JSDAParseError:
            raise
        except Exception as exc:
            raise JSDAParseError(f"xlsxを開けません: {workbook_path}") from exc
        finally:
            if workbook is not None:
                workbook.close()
            if source is not None:
                source.close()
        return

    try:
        import xlrd

        workbook = xlrd.open_workbook(str(workbook_path), on_demand=True)
    except Exception as exc:
        raise JSDAParseError(f"xlsを開けません: {workbook_path}") from exc
    try:
        if sheet_name not in workbook.sheet_names():
            raise JSDAParseError(f"対象シートがありません: {sheet_name}")
        worksheet = workbook.sheet_by_name(sheet_name)
        yield (
            tuple(worksheet.cell_value(row_index, column_index) for column_index in range(worksheet.ncols))
            for row_index in range(worksheet.nrows)
        )
    finally:
        workbook.release_resources()


def _parse_workbook(
    path: str | os.PathLike[str], kind: str, *, min_issue_count: int = 1
) -> dict[str, dict[str, Any]]:
    workbook_path = Path(path)
    if isinstance(min_issue_count, bool) or not isinstance(min_issue_count, int) or min_issue_count < 1:
        raise ValueError("min_issue_countは1以上の整数である必要があります")
    with _workbook_rows(workbook_path, kind) as rows:
        issues: dict[str, dict[str, Any]] = {}
        source_names: dict[str, str] = {}
        bucket_sources: dict[tuple[str, str], set[str]] = {}
        quantity_sums = [0, 0, 0]
        total_quantities: tuple[int, int, int] | None = None
        data_count = 0
        saw_total = False

        for row_number, raw_row in enumerate(rows, 1):
            row = _row_tuple(raw_row)
            if row_number in (6, 7):
                _validate_header(row_number, row, kind)
                continue
            if row_number < 8:
                continue
            if not any(value is not None for value in row):
                continue

            if _normalize_name(row[0]) == "合計":
                if saw_total:
                    raise JSDAParseError("合計行が複数あります")
                if row[1] is not None and str(row[1]).strip():
                    raise JSDAParseError("合計行のコードが空ではありません")
                total_quantities = tuple(
                    _to_int_or_none(
                        row[column],
                        allow_none=False,
                        context=f"合計行 {column + 1}列目",
                    )
                    for column in _QUANTITY_COLUMNS
                )
                saw_total = True
                continue

            if saw_total:
                raise JSDAParseError(f"合計行の後にデータがあります: {row_number}行目")

            source_code = str(row[1]).strip()
            code = _normalize_code(row[1])
            name = _normalize_name(row[0])
            collateral = _COLLATERAL_KEYS.get(row[2])
            if collateral is None:
                raise JSDAParseError(f"{row_number}行目の担保区分が不正です: {row[2]!r}")

            parsed_numbers: dict[int, int | None] = {}
            for column in _NUMERIC_COLUMNS:
                parsed_numbers[column] = _to_int_or_none(
                    row[column],
                    allow_none=column in _CHANGE_COLUMNS,
                    context=f"{row_number}行目 {column + 1}列目",
                )

            measurements = {
                key: parsed_numbers[column] for key, column in _OUTPUT_COLUMNS.items()
            }
            issue = issues.setdefault(code, {"name": name, kind: {}})
            if source_code in source_names and source_names[source_code] != name:
                raise JSDAParseError(
                    f"同一ソースコードの銘柄名が一致しません: {source_code} "
                    f"({source_names[source_code]!r} != {name!r})"
                )
            source_names[source_code] = name

            bucket_key = (code, collateral)
            sources = bucket_sources.setdefault(bucket_key, set())
            if source_code in sources:
                raise JSDAParseError(
                    f"担保区分が重複しています: {source_code} {row[2]}"
                )
            sources.add(source_code)
            if collateral in issue[kind]:
                raise JSDAParseError(
                    f"同一銘柄・同一担保区分の行が重複しています: {code} {row[2]}"
                )
            issue[kind][collateral] = measurements

            for index, column in enumerate(_QUANTITY_COLUMNS):
                value = parsed_numbers[column]
                if not isinstance(value, int):
                    raise JSDAParseError(
                        f"{row_number}行目 {column + 1}列目の数量が整数ではありません: {value!r}"
                    )
                quantity_sums[index] += value
            data_count += 1

        if data_count == 0:
            raise JSDAParseError("データ行が0件です")
        if len(issues) < min_issue_count:
            raise JSDAParseError(
                f"銘柄数が下限未満です: {len(issues)} < {min_issue_count}"
            )
        if total_quantities is None:
            raise JSDAParseError("合計行がありません")
        if tuple(quantity_sums) != total_quantities:
            labels = ("貸付数量", "借入(自己)数量", "借入(転貸)数量")
            differences = ", ".join(
                f"{label}: 明細={actual}, 合計行={expected}"
                for label, actual, expected in zip(labels, quantity_sums, total_quantities)
                if actual != expected
            )
            raise JSDAParseError(f"数量合計が一致しません: {differences}")
        return issues


def parse_zandaka(
    path: str | os.PathLike[str], *, min_issue_count: int = 1
) -> dict[str, dict[str, Any]]:
    """Parse a JSDA weekend-balance workbook into the weekly issues shape."""
    return _parse_workbook(path, "taishaku", min_issue_count=min_issue_count)


def parse_shinki(
    path: str | os.PathLike[str], *, min_issue_count: int = 1
) -> dict[str, dict[str, Any]]:
    """Parse a JSDA weekly-new-contract workbook into the weekly issues shape."""
    return _parse_workbook(path, "shinki", min_issue_count=min_issue_count)


def _validate_source_filename(
    path: str | os.PathLike[str], kind_char: str, report_date: str
) -> None:
    """ファイル名(報告日の正本)が report_date と種別に一致することを検証する。

    命名は ``YYYYMMDD{j,s,z}``。訂正版 ``20260501z(20260514r).xlsx`` や
    フィクスチャの ``20260710z_sample.xlsx`` のような接尾辞は許容する。
    """
    name = Path(path).name
    match = re.match(r"^(\d{8})([jsz])", name)
    if match is None:
        raise JSDAParseError(f"ファイル名がJSDA命名規則(YYYYMMDD[jsz])ではありません: {name}")
    if match.group(2) != kind_char:
        raise JSDAParseError(
            f"ファイル種別が一致しません: {name} (期待: {kind_char})"
        )
    file_date = f"{match.group(1)[:4]}-{match.group(1)[4:6]}-{match.group(1)[6:]}"
    if file_date != report_date:
        raise JSDAParseError(
            f"ファイル名の日付とreport_dateが一致しません: {name} != {report_date}"
        )


def build_weekly(
    z_path: str | os.PathLike[str],
    s_path: str | os.PathLike[str],
    report_date: str,
    *,
    min_issue_count: int = 1,
) -> dict[str, Any]:
    """Build the complete ``supply_demand_weekly_v1`` JSON-compatible object.

    ``min_issue_count`` only gates the z(貸借残高)側(build_z_weeklyと同じ意味の
    下限)。s(新規成約高)は元々z全銘柄より少ない母数が正常(下記コメント)なので、
    同じ下限を課さずparse_shinki側は既定の1のままにする(増分11.5)。
    """
    try:
        parsed_date = date.fromisoformat(report_date)
    except ValueError as exc:
        raise JSDAParseError(f"report_dateがYYYY-MM-DD形式ではありません: {report_date!r}") from exc
    if parsed_date.isoformat() != report_date:
        raise JSDAParseError(f"report_dateがYYYY-MM-DD形式ではありません: {report_date!r}")

    _validate_source_filename(z_path, "z", report_date)
    _validate_source_filename(s_path, "s", report_date)
    zandaka = parse_zandaka(z_path, min_issue_count=min_issue_count)
    # s(新規成約高)は当該週に成約があった銘柄のみ収録される(実測: z=4,330銘柄
    # に対しs=約3,960銘柄)。z/sの銘柄集合は一致しないのが正常で、欠けた側は
    # 空オブジェクトのまま和集合で出力する(仕様)
    shinki = parse_shinki(s_path)
    issues: dict[str, dict[str, Any]] = {}
    for code in sorted(set(zandaka) | set(shinki)):
        z_issue = zandaka.get(code)
        s_issue = shinki.get(code)
        if z_issue and s_issue and z_issue["name"] != s_issue["name"]:
            raise JSDAParseError(
                f"z/sで銘柄名が一致しません: {code} "
                f"({z_issue['name']!r} != {s_issue['name']!r})"
            )
        name = z_issue["name"] if z_issue else s_issue["name"]
        issues[code] = {
            "name": name,
            "taishaku": z_issue["taishaku"] if z_issue else {},
            "shinki": s_issue["shinki"] if s_issue else {},
        }

    return {
        "schema_version": SCHEMA_VERSION,
        "report_date": report_date,
        "source_files": [Path(z_path).name, Path(s_path).name],
        "issues": issues,
    }


def build_z_weekly(
    z_path: str | os.PathLike[str],
    report_date: str,
    *,
    min_issue_count: int = 1,
) -> dict[str, Any]:
    """Build the weekly contract from a z workbook only (backfill v1 scope)."""
    try:
        parsed_date = date.fromisoformat(report_date)
    except ValueError as exc:
        raise JSDAParseError(f"report_dateがYYYY-MM-DD形式ではありません: {report_date!r}") from exc
    if parsed_date.isoformat() != report_date:
        raise JSDAParseError(f"report_dateがYYYY-MM-DD形式ではありません: {report_date!r}")

    _validate_source_filename(z_path, "z", report_date)
    zandaka = parse_zandaka(z_path, min_issue_count=min_issue_count)
    return {
        "schema_version": SCHEMA_VERSION,
        "report_date": report_date,
        "source_files": [Path(z_path).name],
        "issues": {
            code: {
                "name": issue["name"],
                "taishaku": issue["taishaku"],
                "shinki": {},
            }
            for code, issue in sorted(zandaka.items())
        },
    }


def fetch_file(url: str, dest: str | os.PathLike[str]) -> Path:
    """Fetch one file with an explicit UA, two retries, and five-second gaps.

    注意: JSDAは連続リクエストで接続を拒否する(2026-07-22実測)。複数ファイルを
    取得する呼び出し側が、成功時にもファイル間で数秒の間隔を空ける責務を持つ。
    """
    destination = Path(dest)
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(destination.name + ".tmp")
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            request = Request(url, headers={"User-Agent": USER_AGENT})
            with urlopen(request, timeout=60) as response, temporary.open("wb") as output:
                shutil.copyfileobj(response, output)
            os.replace(temporary, destination)
            return destination
        except Exception as exc:
            last_error = exc
            try:
                temporary.unlink()
            except FileNotFoundError:
                pass
            if attempt < 2:
                time.sleep(5)
    raise RuntimeError(f"取得に3回失敗しました: {url}") from last_error


def _main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="JSDA週次xlsxを正規化JSONへ変換します")
    parser.add_argument("z_path", help="銘柄別株券等貸借週末残高 xlsx")
    parser.add_argument("s_path", help="銘柄別株券等貸借週間新規成約高 xlsx")
    parser.add_argument("report_date", help="報告日 (YYYY-MM-DD)")
    args = parser.parse_args(argv)

    result = build_weekly(args.z_path, args.s_path, args.report_date)
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    json.dump(result, sys.stdout, ensure_ascii=False, separators=(",", ":"))
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
